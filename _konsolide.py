# -*- coding: utf-8 -*-
"""C + reverify sonuçlarını birleştir, Baran için tek temiz dosya üret."""
import re, sys
sys.stdout.reconfigure(encoding="utf-8")

def bolumler(dosya):
    txt = open(dosya, encoding="utf-8").read()
    out = {}
    cur = None
    for ln in txt.splitlines():
        h = re.match(r"---\s*(SERBEST|KULÜP DEĞİŞMİŞ|EMEKLİ|AYNI|BULUNAMADI|SD'DE BULUNAMADI)", ln)
        if h:
            cur = h.group(1).replace("SD'DE ", ""); out.setdefault(cur, []); continue
        if cur and ln.strip().startswith(("  ", "\t")) is False and ln.startswith("  "):
            out[cur].append(ln.strip())
    return out

def parse_line(ln):
    # "isim | biz -> kulup" veya "isim | biz"
    m = re.match(r"(.+?) \| (.+?) -> (.+?)(?:\s+\(SD:.*)?$", ln)
    if m: return m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
    m = re.match(r"(.+?) \| (.+?)(?:\s+\(SD:.*)?$", ln)
    if m: return m.group(1).strip(), m.group(2).strip(), ""
    return ln.strip(), "", ""

c = bolumler("_sd_yeniden_ara.txt")
rv = bolumler("_reverify_v2.txt")

serbest, degismis, emekli = {}, {}, {}
for src in (c, rv):
    for ln in src.get("SERBEST", []):
        i,b,_ = parse_line(ln); serbest[i]=b
    for ln in src.get("KULÜP DEĞİŞMİŞ", []):
        i,b,k = parse_line(ln); degismis[i]=(b,k)
    for ln in src.get("EMEKLİ", []):
        i,b,_ = parse_line(ln); emekli[i]=b

# bulunamadı (her iki pas) -> v2'deki tahmini bul (kontrol için)
bulunamadi = set()
for src in (c, rv):
    for ln in src.get("BULUNAMADI", []):
        i,b,_ = parse_line(ln); bulunamadi.add(i)
v2 = open("_sd_kulup_v2.txt", encoding="utf-8").read()
v2_tahmin = {}
for m in re.finditer(r"^\s+(.+?) \| (.+?) -> (.+?)$", v2, re.M):
    v2_tahmin[m.group(1).strip()] = (m.group(2).strip(), m.group(3).strip())
kontrol = {i: v2_tahmin.get(i, ("","?")) for i in bulunamadi if i not in serbest and i not in degismis and i not in emekli}

L = ["=== GÜNCEL KULÜP — BARAN İÇİN (SoccerDonna doğrulaması) ===",
     f"SERBEST kalmış: {len(serbest)} | Kulüp değişmiş: {len(degismis)} | Emekli: {len(emekli)} | Kontrol gerek: {len(kontrol)}",
     "Not: 'Değişmiş' ve 'Serbest' isim+uyruk ile teyitli. 'Kontrol' = SD aramasıyla kesinleşmedi, göz at.",
     f"\n########## SERBEST KALMIŞ ({len(serbest)}) ##########"]
L += [f"  {i}  |  bizde: {b}  →  SERBEST" for i,b in sorted(serbest.items())]
L += [f"\n########## KULÜP DEĞİŞMİŞ ({len(degismis)}) ##########"]
L += [f"  {i}  |  {b}  →  {k}" for i,(b,k) in sorted(degismis.items())]
L += [f"\n########## EMEKLİ ({len(emekli)}) ##########"]
L += [f"  {i}  |  bizde: {b}" for i,b in sorted(emekli.items())]
L += [f"\n########## KONTROL GEREK (SD aramada netleşmedi) ({len(kontrol)}) ##########"]
L += [f"  {i}  |  bizde: {b}  →  (v2 tahmin: {k})" for i,(b,k) in sorted(kontrol.items())]
open("KULUP_GUNCELLEME_BARAN.txt","w",encoding="utf-8").write("\n".join(L))
print("\n".join(L[:3]))
print("\n-> KULUP_GUNCELLEME_BARAN.txt")

# --- Excel (Baran için, Desktop'a) ---
import pandas as pd
from openpyxl.styles import Font, PatternFill
rows = []
for i,b in sorted(serbest.items()):      rows.append([i,b,"SERBEST","🆓 Serbest kalmış"])
for i,(b,k) in sorted(degismis.items()): rows.append([i,b,k,"🔁 Kulüp değişmiş"])
for i,b in sorted(emekli.items()):       rows.append([i,b,"","🏁 Emekli"])
for i,(b,k) in sorted(kontrol.items()):  rows.append([i,b,(k or ""),"🔎 Kontrol gerek"])
df = pd.DataFrame(rows, columns=["İsim","Mevcut Kulüp (bizde)","SD Güncel Kulüp","Durum"])
yol = r"C:\Users\MSI\Desktop\Kulup_Guncelleme.xlsx"
with pd.ExcelWriter(yol, engine="openpyxl") as w:
    df.to_excel(w, index=False, sheet_name="Kulüp Güncelleme")
    ws = w.sheets["Kulüp Güncelleme"]
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for col, wid in zip("ABCD",(30,30,30,18)): ws.column_dimensions[col].width = wid
    fill = PatternFill("solid", fgColor="7C3AED")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = fill
print("XLSX:", yol, "| satır:", len(df))
